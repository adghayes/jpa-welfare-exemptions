"""Render the dataset as a color-coded review workbook (.xlsx).

Every cell whose value came from a manual source is filled with that source's
color; script-generated cells have no fill. A Legend tab documents the scheme.
The xlsx is a *render* of the repo's provenance data — the repo CSVs remain
the source of truth; uploading the xlsx to Google Drive (with conversion)
yields a native, color-coded Google Sheet for review.

Feeds on output/dataset/{grants,parcels,qa_findings}.csv (the merged,
provenance-stamped dataset) and writes output/dataset/review.xlsx with
Grants, Parcels, QA findings, and Legend tabs.

Usage:
    python scripts/publish_review_sheet.py [-o output/dataset/review.xlsx]
"""

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# provenance source -> fill color. Generated/computed cells are filled (very
# lightly); manual cells are UNFILLED — so any column or value a reviewer adds
# directly in the sheet defaults, correctly, to "manual".
GENERATED_FILL = "EAF1EA"             # very light green — script-generated
SOURCE_FILLS = {
    "cmfa-meeting-docs": GENERATED_FILL,
    "cscda-agenda": GENERATED_FILL,
    "cscda-agenda+packet": GENERATED_FILL,
    "la-county-api": GENERATED_FILL,
    "solano-county-portal": GENERATED_FILL,
    "sandag-parcels": GENERATED_FILL,
    "collaborator-sheet": None,       # manual: no fill
    "manual-repo-edit": None,         # manual: no fill
}

LEGEND = [
    ("filled = automated", "Value obtained by automation (document scraping, county API calls) or machine-verified against county records — reproducible by re-running the pipeline"),
    ("no fill = manual", "Human-entered: collaborator sheet research, repo manual/ overrides, sheet formulas — and anything added directly in this sheet"),
]
LEGEND_FILLS = {"filled = automated": GENERATED_FILL, "no fill = manual": None}

HEADER_FILL = PatternFill("solid", fgColor="2F3B33")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)


def render_workbook(tables: dict, out_path: Path) -> None:
    """Write {tab_name: (df, cell_sources)} to out_path with provenance fills.

    cell_sources: {(row_index, column_name): source_string}; sources present
    in SOURCE_FILLS with a color get filled, others stay unfilled (= manual).
    """
    wb = Workbook()
    wb.remove(wb.active)

    for name, (df, cell_sources) in tables.items():
        ws = wb.create_sheet(name)
        for c, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=c, value=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center")
            q90 = df[col].astype(str).str.len().quantile(0.9) if len(df) else 12
            width = max(12, min(38, int(q90) + 2))
            ws.column_dimensions[get_column_letter(c)].width = width
        ws.freeze_panes = "C2"

        for r, (idx, row) in enumerate(df.iterrows(), start=2):
            for c, col in enumerate(df.columns, start=1):
                cell = ws.cell(row=r, column=c, value=row[col] if row[col] != "" else None)
                if row[col] == "":
                    continue
                if col.endswith("_url"):
                    # render as a click-through link named by the filename
                    cell.value = row[col].rstrip("/").split("/")[-1]
                    cell.hyperlink = row[col]
                    cell.font = Font(color="1155CC", underline="single", size=10)
                fill = SOURCE_FILLS.get(cell_sources.get((idx, col)))
                if fill:
                    cell.fill = PatternFill("solid", fgColor=fill)

    legend = wb.create_sheet("Legend")
    legend.column_dimensions["A"].width = 22
    legend.column_dimensions["B"].width = 110
    legend.cell(row=1, column=1, value="Cell color").font = Font(bold=True)
    legend.cell(row=1, column=2, value="Meaning").font = Font(bold=True)
    for i, (source, meaning) in enumerate(LEGEND, start=2):
        c = legend.cell(row=i, column=1, value=source)
        fill = LEGEND_FILLS.get(source)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        legend.cell(row=i, column=2, value=meaning)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"wrote {out_path} ({out_path.stat().st_size:,} bytes)")


# ---------------------------------------------------------- dataset feed

# columns that are identity/metadata, never colored as "automated"
GRANT_META = {"project_id", "row_source", "field_overrides", "manual_status",
              "superseded_by"}
PARCEL_IDENTITY = {"project_id", "operative_project_id", "county",
                   "property_name", "ain", "apn", "situs_address",
                   "legacy_redundant", "assignment_source", "method", "notes"}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("-o", "--output", type=Path,
                        default=Path("output/dataset/review.xlsx"))
    args = parser.parse_args()

    grants = pd.read_csv("output/dataset/grants.csv", dtype=str).fillna("")
    parcels = pd.read_csv("output/dataset/parcels.csv", dtype=str).fillna("")

    # grants: generated cells filled, manually-overridden fields unfilled
    grant_sources = {}
    for idx, row in grants.iterrows():
        overridden = {e.split(":")[0] for e in row["field_overrides"].split("; ") if e}
        fully_manual = row["row_source"] != "generated" or "*" in overridden
        for col in grants.columns:
            if row[col] == "":
                continue
            if col in ("authorization_status", "jpa_closing_fee",
                       "jpa_annual_fee", "years_since_approval"):
                # fully calculated (minutes rollup / published fee schedules)
                # even on otherwise-manual rows — always tinted as automated
                grant_sources[(idx, col)] = "cmfa-meeting-docs"
            elif fully_manual or col in GRANT_META or col in overridden:
                grant_sources[(idx, col)] = "collaborator-sheet"   # unfilled
            else:
                grant_sources[(idx, col)] = "cmfa-meeting-docs"  # any generated fill

    # parcels: identity columns are manual assignments (unfilled); value
    # columns are filled only when they came from the county API
    parcel_sources = {}
    MANUAL_SOURCES = {"", "collaborator-sheet", "manual-repo-edit"}
    for idx, row in parcels.iterrows():
        api = row["value_source"] not in MANUAL_SOURCES
        validated = row.get("assignment_check", "") == "situs-match"
        for col in parcels.columns:
            if row[col] == "":
                continue
            if col in ("ain", "apn") and validated:
                # identity is a human assertion, but this one is machine-
                # verified against the county situs — tint it
                parcel_sources[(idx, col)] = "la-county-api"
            elif col == "assignment_check":
                parcel_sources[(idx, col)] = "la-county-api"  # computed
            elif col in PARCEL_IDENTITY or not api:
                parcel_sources[(idx, col)] = "collaborator-sheet"  # unfilled
            else:
                parcel_sources[(idx, col)] = "la-county-api"

    # QA findings: merge-time discrepancies for collaborator review
    # (document-vs-manual conflicts, grants missing from either side,
    # shared AINs, parcels without values). Meta table — no fills.
    qa = pd.read_csv("output/dataset/qa_findings.csv", dtype=str).fillna("")

    render_workbook({"Grants": (grants, grant_sources),
                     "Parcels": (parcels, parcel_sources),
                     "QA findings": (qa, {})}, args.output)


if __name__ == "__main__":
    main()
